from robot.libraries.BuiltIn import BuiltIn
import openpyxl
wk = openpyxl.load_workbook("testData\\dataRobot.xlsx")

def recuperer_max_ligne(sheetName):
    sh = wk[sheetName]
    return sh.max_row

#print(recuperer_max_ligne("smoke"))

def recuperer_max_colonne(sheetName):
    sh = wk[sheetName]
    return sh.max_column
#print(recuperer_max_colonne("smoke"))

def recuperer_donnees_cellules(sheetname,row,cell):
    sh = wk[sheetname]
    cellule = sh.cell(int(row),int(cell))
    return cellule.value 
#print(recuperer_donnees_cellules("data",4,2))
# seleniumLB = BuiltIn().get_library_instance("SeleniumLibrary")
# allureLB = BuiltIn().get_library_instance("AllureLibrary")
# def capture_screenshot():
#     filemame = "screen.png"
#     seleniumLB.capture_page_screenshot(filemame)
#     allureLB.attach_file(f'./{filemame}')